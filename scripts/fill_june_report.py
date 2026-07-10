"""KHIDI 6월 월간 업무 보고 자동 채움 (어시스턴트 초안 2026-07-10).

- 초안 본문: docs/government-project/monthly-reports/2026-06_월간보고_초안.md
- 형식: scripts/fill_may_report.py 와 동일 (6월 시트 C=6월 실적 / D=7월 계획)
- KPI(C9~C11)는 2026-07-10 실DB 실측 — 테스트·데모 시드 전면 제외 기준.
- [PO 확인] 표시 셀은 제출 전 PO가 채우거나 지울 것.
"""
from openpyxl import load_workbook
import shutil
import os

SRC = r"C:\Users\user\Documents\테플러\2025 정부지원과제\02. 본로이\03. 진행 중\12. ICT 기반 외국인환자 사전상담·사후관리 지원 사업\02. 진행서류\07. 월간 업무 보고\월간 업무 보고_5월_본로이_작성본.xlsx"
DST = r"C:\Users\user\Documents\테플러\2025 정부지원과제\02. 본로이\03. 진행 중\12. ICT 기반 외국인환자 사전상담·사후관리 지원 사업\02. 진행서류\07. 월간 업무 보고\월간 업무 보고_6월_본로이_작성본.xlsx"

shutil.copy2(SRC, DST)

# 6월 추진 실적 (C5~C8)
JUNE_C = {
    5: """○ 웹사이트 개발
 - 성과지표 추적 대시보드 구현 (유치 전환 현황, 기관별 집계)
 - 양·한방 협진 의뢰 워크플로우 구축 (협진 연계율 측정)

○ 글로벌 외국인 환자 유치 에이전시 발굴
 - 에이전시 전용 포털 구축 (환자 케이스 진행상황·보험정보 공유)""",
    6: """○ 웹사이트 개발
 - 상담기록·협진의뢰서 증빙자료 내보내기 기능 구현
 - 성과지표 집계 정확성 정비 및 집계오류 자동 감시 체계 적용

○ 글로벌 외국인 환자 유치 에이전시 발굴
 - 해외 의료기관 치료경과 업로드 기능 구현 (현지 주치의 사후관리 연계)""",
    7: """○ 웹사이트 개발
 - 시스템 보안 전수 점검 및 보완 (고위험 취약점 0건 확인)
 - 환자 만족도 설문 자동발송 체계 정비
 - AI 상담 품질관리·안전장치 강화 (부정확한 표현 자동 차단, 6개 언어)""",
    8: """○ 웹사이트 개발
 - AI 상담 → 코디네이터 연결 동선 개선 (예시질문·연결버튼, 6개 언어)
 - 관리자 성과지표 통합 현황판 구축 (목표 대비 달성률 실시간 확인)

○ 글로벌 외국인 환자 유치 에이전시 발굴
 - [PO 확인] 6월 에이전시 미팅·MOU 등 오프라인 활동""",
}

# 7월 추진 계획 (D5~D8)
JUNE_D = {
    5: """○ 웹사이트 개발
 - 견적서·동의서·비자초청장 등 발급문서 다국어(러시아·카자흐어) 품질 정비

○ 글로벌 외국인 환자 유치 에이전시 발굴
 - 보험사 제휴 채널 신설 (러시아 보험·어시스턴스사 제휴 추진, 보험 안내 페이지 6개 언어)""",
    6: """○ 웹사이트 개발
 - 검색 노출 최적화 (네이버·구글 검색 등록 점검)

○ 글로벌 외국인 환자 유치 에이전시 발굴
 - CIS권 에이전시 추가 미팅 및 플랫폼 가입 유도 (키르기스스탄 등)""",
    7: """○ 웹사이트 개발
 - AI 상담 품질 상시 관측 강화 (정확도·응답시간 자동 채점)
 - 참여병원 4개 지점 외국인환자 유치의료기관 등록 반영·의료진 정보 갱신""",
    8: """○ 글로벌 외국인 환자 유치 에이전시 발굴
 - 실환자 사전상담 → 유치 전환 개시 (에이전시·보험사 경유 실환자 유입)
 - 8월 중간평가 대비 성과 정리 및 발표자료 준비 착수""",
}

# 6월 KPI 실측 (2026-07-10 실DB, 테스트·데모 제외)
JUNE_KPI = {9: 0, 10: 0, 11: 0}
# 7월 계획 목표 — 제안값 [PO 확인 후 확정]
JUNE_KPI_PLAN = {9: 3, 10: 0, 11: 1}

JUNE_ETC = """• 6월 신규 외국인환자 문의 6건 접수 (실데이터 기준) — 사전상담 전환은 7월부터 본격화
• 성과지표 집계 기준 정비: 테스트·시연용 데이터 전면 제외, 실환자 실적만 집계 (보고 수치 신뢰성 확보)
• 시스템 보안 전수 점검 완료 — 고위험 취약점 0건
• 플랫폼 개발 마일스톤(6~7월 개발 완료) 정상 진행: 원격협진 안정화·6개 언어 지원·성과지표 자동 집계 체계 완비
• [PO 확인] 6월 사업비 집행액 (요약 시트 월별 사용액 칸)"""

wb = load_workbook(DST)
ws = wb["6월"]

for r, t in JUNE_C.items():
    c = ws.cell(row=r, column=3)
    c.value = t
    c.alignment = c.alignment.copy(wrap_text=True, vertical="top")

for r, t in JUNE_D.items():
    c = ws.cell(row=r, column=4)
    c.value = t
    c.alignment = c.alignment.copy(wrap_text=True, vertical="top")

for r, v in JUNE_KPI.items():
    ws.cell(row=r, column=3).value = v
for r, v in JUNE_KPI_PLAN.items():
    ws.cell(row=r, column=4).value = v

ws.cell(row=12, column=3).value = JUNE_ETC
c12 = ws.cell(row=12, column=3)
c12.alignment = c12.alignment.copy(wrap_text=True, vertical="top")

for r in [5, 6, 7, 8]:
    ws.row_dimensions[r].height = 160
ws.row_dimensions[12].height = 110

wb.save(DST)
print("OK saved:", DST)
print("size:", os.path.getsize(DST), "bytes")

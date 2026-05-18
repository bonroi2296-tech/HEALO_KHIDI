"""KHIDI 5월 월간 업무 보고 자동 채움 (사용자 초안 기반)."""
from openpyxl import load_workbook
import shutil
import os

SRC = r"C:\Users\user\Documents\테플러\2025 정부지원과제\02. 본로이\03. 진행 중\12. ICT 기반 외국인환자 사전상담·사후관리 지원 사업\02. 진행서류\07. 월간 업무 보고\월간 업무 보고_5월_본로이.xlsx"
DST = r"C:\Users\user\Documents\테플러\2025 정부지원과제\02. 본로이\03. 진행 중\12. ICT 기반 외국인환자 사전상담·사후관리 지원 사업\02. 진행서류\07. 월간 업무 보고\월간 업무 보고_5월_본로이_작성본.xlsx"

shutil.copy2(SRC, DST)

# 사용자 초안 그대로
APRIL_C = {
    5: """○ 웹사이트 개발
 - 기존 개발 웹사이트(HEALO) 재설계
 - 카자흐, 러시아어 적용

○ 글로벌 외국인 환자 유치 에이전시 발굴
 - MOU 계약서 초안 구성
 - Medyvoyage 업체 미팅""",
    6: """○ 웹사이트 개발
 - 홈페이지 UI/UX 개선
 - 환자 문의 접수 페이지 개선

○ 글로벌 외국인 환자 유치 에이전시 발굴
 - 에이전시 리스트 리서치 및 적합성 분류
 - MOU 계약서 완성""",
    7: """○ 웹사이트 개발
 - AI 챗봇 고도화
 - 환자용 대시보드 설계 및 적용

○ 글로벌 외국인 환자 유치 에이전시 발굴
 - Global Health Opulence 업체 미팅
 - The Medical Travel Company 업체 미팅""",
    8: """○ 웹사이트 개발
 - 원격 화상상담 설계 및 적용
 - 시스템 보안 강화

○ 글로벌 외국인 환자 유치 에이전시 발굴
 - Medyonix 업체 미팅
 - Global Healthcare Opulence MOU 체결 완료
 - 카자흐 국적 코디네이터 면접 진행""",
}

# 5월 계획 — 사용자 작성분 + 비어있던 3주·4주 에이전시 섹션 추가
APRIL_D = {
    5: """○ 웹사이트 개발
 - 데이터 백업 이중화
 - SEO 검색 최적화

○ 글로벌 외국인 환자 유치 에이전시 발굴
 - Medyvoyage MOU 체결
 - 에이전시 추가 발굴""",
    6: """○ 웹사이트 개발
 - 원격 화상상담 실시간 자막 적용
 - 환자 만족도 설문 시스템 적용

○ 글로벌 외국인 환자 유치 에이전시 발굴
 - MOU 에이전시 등록 외국인 환자 적합성 검토""",
    7: """○ 웹사이트 개발
 - 성과 지표 추적 대시보드 구현
 - 사후관리 지원 시스템 설계

○ 글로벌 외국인 환자 유치 에이전시 발굴
 - 에이전시별 환자 송출 프로세스 협의
 - 카자흐·러시아 종양내과 의사 1차 컨택 (LinkedIn·Telegram)
 - 신규 에이전시 추가 미팅 (CIS·중동권)""",
    8: """○ 웹사이트 개발
 - 카자흐, 러시아어 전수 검수
 - 시범 운영 시작

○ 글로벌 외국인 환자 유치 에이전시 발굴
 - 첫 외국인 환자 유치 시범 진행 (1~2건)
 - 에이전시 피드백 수집 및 시스템 개선
 - 6월 추가 에이전시 발굴 리스트 확정""",
}

APRIL_KPI = {9: 0, 10: 0, 11: 0}
APRIL_KPI_PLAN = {9: 0, 10: 0, 11: 0}  # 5월 시범 운영 단계, 본격 환자 6월부터

APRIL_ETC = """• 4월 30일 기준 시스템 구축 완료도 약 74%
• 면력한방병원 4개 지점 의료진·치료법 데이터 통합 완료
• 6개국어 UI 지원 (구글·Yandex 검색 노출 준비 완료)
• Global Healthcare Opulence MOU 체결 완료, 추가 에이전시 협상 진행 중
• 카자흐 국적 코디네이터 면접 진행 중
• 5월 4주차 시범 운영 후 6월부터 본격 환자 유치"""

wb = load_workbook(DST)
ws = wb["4월"]

for r, t in APRIL_C.items():
    c = ws.cell(row=r, column=3)
    c.value = t
    c.alignment = c.alignment.copy(wrap_text=True, vertical="top")

for r, t in APRIL_D.items():
    c = ws.cell(row=r, column=4)
    c.value = t
    c.alignment = c.alignment.copy(wrap_text=True, vertical="top")

for r, v in APRIL_KPI.items():
    ws.cell(row=r, column=3).value = v
for r, v in APRIL_KPI_PLAN.items():
    ws.cell(row=r, column=4).value = v

ws.cell(row=12, column=3).value = APRIL_ETC
c12 = ws.cell(row=12, column=3)
c12.alignment = c12.alignment.copy(wrap_text=True, vertical="top")

for r in [5, 6, 7, 8]:
    ws.row_dimensions[r].height = 160
ws.row_dimensions[12].height = 110

wb.save(DST)
print("OK saved:", DST)
print("size:", os.path.getsize(DST), "bytes")
